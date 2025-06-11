import tkinter as tk
from tkinter import filedialog, ttk, messagebox, scrolledtext
import pandas as pd
import json
import os
import time
import threading
import traceback
import csv
import asyncio
import pickle
from datetime import datetime
from typing import Dict, List, Any, Tuple
import aiohttp
import tenacity
import logging
from openai import APIError, RateLimitError, APIConnectionError, APITimeoutError

# Excel出力のためのライブラリを追加
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 新しいOpenAI SDKに対応
from openai import AsyncOpenAI

class FlagAnalysisApp:
    def __init__(self, root):
        self.root = root
        self.root.title("議員発言フラグ分析ツール（改良版）")
        self.root.geometry("1200x800")
        
        # データ保持用変数
        self.texts_df = None
        self.flags_df = None
        self.results_df = None
        self.token_usage_df = None  # トークン使用量データフレーム
        self.checkpoint_file = "analysis_checkpoint.pkl"
        self.batch_size = 1  # バッチサイズをデフォルト1に変更
        
        # 処理状態管理
        self.processed_indices = set()
        self.total_processed = 0
        self.error_log = []  # エラーログ
        self.failed_indices = set()  # 処理失敗したインデックス
        self.is_processing = False
        self.stop_requested = False  # 停止要求フラグ
        self.rate_limit_count = 0  # レート制限エラー回数
        self.dynamic_wait_time = 3.0  # 動的待機時間をデフォルト3秒に変更
        
        # ログ設定
        self.setup_logging()
        
        self.setup_ui()
    
    def setup_logging(self):
        """ログ設定を初期化"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('flag_analysis.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def setup_ui(self):
        # メインフレーム
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # ファイル選択セクション
        file_frame = ttk.LabelFrame(main_frame, text="ファイル選択", padding=10)
        file_frame.pack(fill=tk.X, pady=5)
        
        # 議員発言ファイル選択
        ttk.Label(file_frame, text="議員発言ファイル (testtext.csv):").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.text_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.text_file_var, width=50).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="参照...", command=self.browse_text_file).grid(row=0, column=2, padx=5, pady=5)
        
        # フラグファイル選択
        ttk.Label(file_frame, text="フラグ定義ファイル (huragu.csv):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.flag_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.flag_file_var, width=50).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="参照...", command=self.browse_flag_file).grid(row=1, column=2, padx=5, pady=5)
        
        # 設定セクション
        settings_frame = ttk.LabelFrame(main_frame, text="API設定", padding=10)
        settings_frame.pack(fill=tk.X, pady=5)
        
        # 基本設定
        ttk.Label(settings_frame, text="Temperature:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.temp_var = tk.DoubleVar(value=0.05)  # デフォルト値を0.05に変更
        temp_spin = ttk.Spinbox(settings_frame, from_=0.0, to=1.0, increment=0.01, textvariable=self.temp_var, width=5)
        temp_spin.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        # バッチサイズ設定
        ttk.Label(settings_frame, text="バッチサイズ:").grid(row=0, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        self.batch_size_var = tk.IntVar(value=1)  # デフォルト値を1に変更
        batch_spin = ttk.Spinbox(settings_frame, from_=1, to=10, textvariable=self.batch_size_var, width=5)
        batch_spin.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        
        # 出力形式選択
        ttk.Label(settings_frame, text="出力形式:").grid(row=0, column=4, sticky=tk.W, pady=5, padx=(20, 0))
        self.output_format_var = tk.StringVar(value="excel")
        excel_radio = ttk.Radiobutton(settings_frame, text="Excel", variable=self.output_format_var, value="excel")
        excel_radio.grid(row=0, column=5, sticky=tk.W, padx=5, pady=5)
        csv_radio = ttk.Radiobutton(settings_frame, text="CSV", variable=self.output_format_var, value="csv")
        csv_radio.grid(row=0, column=6, sticky=tk.W, padx=5, pady=5)
        
        # チェックポイント機能
        checkpoint_frame = ttk.LabelFrame(main_frame, text="進捗管理", padding=10)
        checkpoint_frame.pack(fill=tk.X, pady=5)
        
        self.checkpoint_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(checkpoint_frame, text="チェックポイント機能を有効にする", variable=self.checkpoint_var).pack(side=tk.LEFT)
        ttk.Button(checkpoint_frame, text="チェックポイントをリセット", command=self.reset_checkpoint).pack(side=tk.LEFT, padx=10)
        
        # 失敗アイテム再処理オプション
        self.retry_failed_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(checkpoint_frame, text="失敗したアイテムを再処理する", variable=self.retry_failed_var).pack(side=tk.LEFT, padx=10)
        
        # 実行ボタン
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(button_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.analyze_button = ttk.Button(button_frame, text="フラグ分析実行", command=self.start_analysis)
        self.analyze_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = ttk.Button(button_frame, text="停止", command=self.stop_analysis, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        self.save_button = ttk.Button(button_frame, text="結果を保存", command=self.save_results, state=tk.DISABLED)
        self.save_button.pack(side=tk.LEFT, padx=5)
        
        self.error_report_button = ttk.Button(button_frame, text="エラーレポート", command=self.show_error_report, state=tk.DISABLED)
        self.error_report_button.pack(side=tk.LEFT, padx=5)
        
        # ログと結果表示エリア
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # ログタブ
        log_frame = ttk.Frame(notebook)
        notebook.add(log_frame, text="ログ")
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 結果タブ
        result_frame = ttk.Frame(notebook)
        notebook.add(result_frame, text="結果")
        
        # 結果表示用のツリービュー（横スクロール可能にする）
        result_container = ttk.Frame(result_frame)
        result_container.pack(fill=tk.BOTH, expand=True)
        
        # 垂直スクロールバー
        y_scrollbar = ttk.Scrollbar(result_container, orient="vertical")
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 水平スクロールバー
        x_scrollbar = ttk.Scrollbar(result_container, orient="horizontal")
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.result_tree = ttk.Treeview(result_container, yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        self.result_tree.pack(fill=tk.BOTH, expand=True)
        
        y_scrollbar.config(command=self.result_tree.yview)
        x_scrollbar.config(command=self.result_tree.xview)
        
        # ステータスバー
        self.status_var = tk.StringVar()
        self.status_var.set("準備完了")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def browse_text_file(self):
        file_path = filedialog.askopenfilename(
            title="議員発言ファイルを選択",
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
        )
        if file_path:
            self.text_file_var.set(file_path)
    
    def browse_flag_file(self):
        file_path = filedialog.askopenfilename(
            title="フラグ定義ファイルを選択",
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
        )
        if file_path:
            self.flag_file_var.set(file_path)
    
    def log(self, message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
        # ファイルにもログ出力
        self.logger.info(message)
    
    def stop_analysis(self):
        """分析処理を停止"""
        self.stop_requested = True
        self.log("停止要求を受け付けました...")
        self.status_var.set("停止中...")
    
    def show_error_report(self):
        """エラーレポートを表示"""
        if not self.error_log:
            messagebox.showinfo("エラーレポート", "エラーはありませんでした")
            return
        
        error_window = tk.Toplevel(self.root)
        error_window.title("エラーレポート")
        error_window.geometry("800x600")
        
        error_text = scrolledtext.ScrolledText(error_window)
        error_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        for error in self.error_log:
            error_text.insert(tk.END, f"{error}\n{'-'*50}\n")
        
        # エラーレポートを保存ボタン
        def save_error_report():
            file_path = filedialog.asksaveasfilename(
                title="エラーレポートを保存",
                defaultextension=".txt",
                filetypes=[("テキストファイル", "*.txt")]
            )
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    for error in self.error_log:
                        f.write(f"{error}\n{'-'*50}\n")
                messagebox.showinfo("保存完了", f"エラーレポートを保存しました: {file_path}")
        
        ttk.Button(error_window, text="保存", command=save_error_report).pack(pady=5)
    
    def reset_checkpoint(self):
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
            self.processed_indices.clear()
            self.failed_indices.clear()
            self.total_processed = 0
            self.error_log.clear()
            if hasattr(self, 'token_usage_df'):
                self.token_usage_df = pd.DataFrame(columns=["名前", "インプットトークン", "アウトプットトークン"])
            self.log("チェックポイントをリセットしました")
            messagebox.showinfo("完了", "チェックポイントをリセットしました")
    
    def load_checkpoint(self):
        """チェックポイントを読み込む"""
        if not self.checkpoint_var.get() or not os.path.exists(self.checkpoint_file):
            return None
        
        try:
            with open(self.checkpoint_file, 'rb') as f:
                checkpoint_data = pickle.load(f)
                self.processed_indices = checkpoint_data.get('processed_indices', set())
                self.failed_indices = checkpoint_data.get('failed_indices', set())
                self.total_processed = checkpoint_data.get('total_processed', 0)
                self.error_log = checkpoint_data.get('error_log', [])
                results_data = checkpoint_data.get('results_df', None)
                if results_data is not None:
                    self.results_df = pd.DataFrame(results_data)
                
                # トークン使用量データの復元
                token_data = checkpoint_data.get('token_usage_df', [])
                if isinstance(token_data, list) and token_data:
                    self.token_usage_df = pd.DataFrame(token_data)
                else:
                    self.token_usage_df = pd.DataFrame(columns=["名前", "インプットトークン", "アウトプットトークン"])
                
                self.log(f"チェックポイントから復元: {self.total_processed}件処理済み, {len(self.failed_indices)}件失敗")
                return checkpoint_data
        except Exception as e:
            self.log(f"チェックポイント読み込みエラー: {e}")
            return None
    
    def save_checkpoint(self):
        """チェックポイントを保存する"""
        if not self.checkpoint_var.get():
            return
        
        try:
            checkpoint_data = {
                'processed_indices': self.processed_indices,
                'failed_indices': self.failed_indices,
                'total_processed': self.total_processed,
                'error_log': self.error_log,
                'results_df': self.results_df.to_dict('records') if self.results_df is not None else None,
                'token_usage_df': self.token_usage_df.to_dict('records') if not self.token_usage_df.empty else [],
                'timestamp': datetime.now().isoformat()
            }
            with open(self.checkpoint_file, 'wb') as f:
                pickle.dump(checkpoint_data, f)
        except Exception as e:
            self.log(f"チェックポイント保存エラー: {e}")
            self.logger.error(f"Checkpoint save error: {e}", exc_info=True)
    
    def start_analysis(self):
        # 非同期で分析を実行
        self.analyze_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.stop_requested = False
        self.is_processing = True
        self.batch_size = self.batch_size_var.get()
        thread = threading.Thread(target=self.run_async_analysis)
        thread.daemon = True
        thread.start()
    
    def run_async_analysis(self):
        """非同期分析をイベントループで実行"""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(self.analyze_flags_async())
        finally:
            loop.close()
            self.is_processing = False
            self.analyze_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.error_report_button.config(state=tk.NORMAL if self.error_log else tk.DISABLED)
    
    async def analyze_flags_async(self):
        try:
            # ファイルの存在確認とデータ読み込み
            text_file = self.text_file_var.get()
            flag_file = self.flag_file_var.get()
            
            if not text_file or not flag_file:
                messagebox.showerror("エラー", "両方のファイルを選択してください")
                return
            
            self.status_var.set("データ読み込み中...")
            self.log(f"発言ファイル読み込み: {text_file}")
            self.log(f"フラグファイル読み込み: {flag_file}")
            
            # CSVファイル読み込み
            try:
                self.texts_df = pd.read_csv(text_file, encoding='utf-8')
                self.flags_df = pd.read_csv(flag_file, encoding='utf-8', dtype=str)
                self.flags_df["小項目コード"] = self.flags_df["小項目コード"].astype(int)
                
                self.log(f"発言データ: {len(self.texts_df)}行")
                self.log(f"フラグデータ: {len(self.flags_df)}行")
            except Exception as e:
                self.log(f"ファイル読み込みエラー: {e}")
                messagebox.showerror("読み込みエラー", f"CSVファイルの読み込みに失敗しました: {e}")
                return
            
            # チェックポイントを読み込み
            self.load_checkpoint()
            
            # 結果を格納するデータフレームを準備
            if self.results_df is None:
                columns = ["名前"]
                for code in self.flags_df["小項目コード"]:
                    code_str = str(code)
                    columns.extend([f"{code_str}_flag", f"{code_str}_quote", f"{code_str}_reason"])
                self.results_df = pd.DataFrame(columns=columns)
            
            # トークン使用量データフレームを準備
            if self.token_usage_df is None:
                self.token_usage_df = pd.DataFrame(columns=["名前", "インプットトークン", "アウトプットトークン"])
            
            # Structured Outputs用のスキーマを定義
            json_schema = self.create_json_schema()
            
            # APIクライアントを初期化
            api_key = os.getenv("OPENAI_API_KEY")
            if not api_key:
                raise ValueError("OpenAI APIキーが設定されていません。環境変数OPENAI_API_KEYを確認してください。")
            
            client = AsyncOpenAI(api_key=api_key)
            
            # プロンプトテンプレートを作成
            prompt_template = self.create_prompt_template()
            
            # バッチ処理で分析実行
            self.log("フラグ分析開始...")
            total = len(self.texts_df)
            
            # 未処理のデータを取得
            unprocessed_indices = []
            for i in range(total):
                if i not in self.processed_indices:
                    # 失敗アイテムの再処理オプションをチェック
                    if i in self.failed_indices and not self.retry_failed_var.get():
                        continue  # 再処理しない
                    unprocessed_indices.append(i)
            
            if not unprocessed_indices:
                self.log("すべてのデータが処理済みです")
                self.display_results()
                self.save_button.config(state=tk.NORMAL)
                return
            
            self.log(f"未処理データ: {len(unprocessed_indices)}件")
            if self.failed_indices and self.retry_failed_var.get():
                retry_count = len([i for i in unprocessed_indices if i in self.failed_indices])
                self.log(f"うち再処理: {retry_count}件")
            
            # バッチごとに処理
            for batch_start in range(0, len(unprocessed_indices), self.batch_size):
                # 停止要求チェック
                if self.stop_requested:
                    self.log("ユーザーによる停止要求を受信しました")
                    break
                
                batch_indices = unprocessed_indices[batch_start:batch_start + self.batch_size]
                batch_data = []
                
                for idx in batch_indices:
                    row = self.texts_df.iloc[idx]
                    batch_data.append((idx, row["名前"], row["テキスト"]))
                
                if not batch_data:
                    continue
                
                # バッチ処理
                try:
                    await self.process_batch(client, batch_data, prompt_template, json_schema)
                    # 成功時：待機時間を少し短縮
                    self.dynamic_wait_time = max(1.5, self.dynamic_wait_time * 0.95)
                except Exception as e:
                    self.log(f"バッチ処理でエラー: {e}")
                    self.logger.error(f"Batch processing error: {e}", exc_info=True)
                    
                    # レート制限エラーの場合は待機時間を大幅に増加
                    if "rate limit" in str(e).lower() or "429" in str(e):
                        self.rate_limit_count += 1
                        self.dynamic_wait_time = min(30.0, self.dynamic_wait_time * 2)
                        self.log(f"レート制限により待機時間を{self.dynamic_wait_time}秒に調整")
                
                # 進捗更新
                self.total_processed = len(self.processed_indices)
                progress = (self.total_processed / total) * 100
                self.progress_var.set(progress)
                self.status_var.set(f"分析中... ({self.total_processed}/{total}, 失敗: {len(self.failed_indices)})")
                
                # チェックポイント保存（バッチ毎）
                self.save_checkpoint()
                
                # 動的レート制限対策（デフォルト3秒に調整）
                actual_wait = max(3.0, self.dynamic_wait_time)
                self.log(f"待機中... ({actual_wait:.1f}秒)")
                await asyncio.sleep(actual_wait)
            
            # 最終結果表示
            self.display_results()
            self.progress_var.set(100)
            
            if self.stop_requested:
                self.status_var.set(f"ユーザーにより停止 - 完了: {self.total_processed}件, 失敗: {len(self.failed_indices)}件")
                messagebox.showwarning("停止", f"処理を停止しました。\n完了: {self.total_processed}件\n失敗: {len(self.failed_indices)}件")
            else:
                self.status_var.set(f"分析完了 - 完了: {self.total_processed}件, 失敗: {len(self.failed_indices)}件")
                messagebox.showinfo("完了", f"フラグ分析が完了しました。\n処理件数: {self.total_processed}\n失敗件数: {len(self.failed_indices)}")
            
            self.save_button.config(state=tk.NORMAL)
            
            # 失敗したアイテムのレポート
            if self.failed_indices:
                failed_names = []
                for idx in self.failed_indices:
                    if idx < len(self.texts_df):
                        failed_names.append(self.texts_df.iloc[idx]["名前"])
                self.log(f"処理に失敗したアイテム: {', '.join(failed_names[:5])}{'...' if len(failed_names) > 5 else ''}")
            
        except Exception as e:
            self.log(f"分析エラー: {e}")
            self.log(f"詳細: {traceback.format_exc()}")
            messagebox.showerror("エラー", f"分析中にエラーが発生しました: {e}")
    
    def create_json_schema(self):
        """Structured Outputs用のJSONスキーマを作成（enum制限対応）"""
        properties = {}
        for code in self.flags_df["小項目コード"]:
            code_str = str(code)
            properties[code_str] = {
                "type": "object",
                "properties": {
                    "flag": {
                        "type": "integer",
                        "minimum": 0,
                        "maximum": 2,
                        "description": "0: 言及無し, 1: 言及されているが公約ではない, 2: 公約として明確に言及"
                    },
                    "quote": {
                        "type": "string",
                        "description": "フラグが1または2の場合の根拠となる発言部分の引用"
                    },
                    "reason": {
                        "type": "string",
                        "description": "フラグが1または2の場合の理由（10-30文字）"
                    }
                },
                "required": ["flag", "quote", "reason"],
                "additionalProperties": False
            }
        
        return {
            "type": "object",
            "properties": properties,
            "required": list(properties.keys()),
            "additionalProperties": False
        }
    
    def create_prompt_template(self):
        """プロンプトテンプレートを作成"""
        prompt_items = []
        for _, row in self.flags_df.iterrows():
            code = row["小項目コード"]
            description = row["説明"]
            prompt_items.append(f"{code}: {description}")
        
        return f"""
以下の政治発言を分析し、各政策テーマに対してフラグを設定してください。

フラグの定義:
- 0: 言及無し
- 1: 言及されているが公約ではない（言及・検討・議論等）
- 2: 公約として明確に言及（実施・実現・推進・実行等の強い表現）

フラグが1または2の場合は、以下を含めてください:
1. 根拠となる発言部分の引用（元の発言からそのまま抜粋すること）
2. フラグが立つ理由（10-30文字で詳細に説明）

重要:
- 引用は必ず元の発言から一字一句そのまま抜粋してください
- 要約や意訳ではなく、実際の発言をそのまま引用してください
- 引用部分が長すぎる場合は、最も重要な部分を抜粋してください

1つの発言が複数のフラグに関連する場合も適切に判定してください。
文脈から公約かどうかを判断し、100%の精度は求めませんが合理的な判定を行ってください。

**重要**: 必ず完全で有効なJSONで応答してください。文字列内の引用符は適切にエスケープしてください。

政策テーマ（コード: 説明）:
{chr(10).join(prompt_items)}

発言: {{text}}
"""
    
    @tenacity.retry(
        stop=tenacity.stop_after_attempt(5),
        wait=tenacity.wait_exponential(multiplier=2, min=5, max=120),  # 最大待機2分に拡張
        retry=tenacity.retry_if_exception_type((
            RateLimitError, 
            APIConnectionError, 
            APITimeoutError,
            ConnectionError,
            TimeoutError
        )),
        before_sleep=lambda retry_state: retry_state.outcome.exception() and 
            logging.getLogger(__name__).warning(f"API呼び出し失敗、{retry_state.next_action}秒後にリトライ (試行{retry_state.attempt_number}/5): {retry_state.outcome.exception()}")
    )
    async def call_openai_api(self, client, prompt, json_schema):
        """OpenAI APIを呼び出す（改良されたレート制限対応）"""
        try:
            response = await client.chat.completions.create(
                model="gpt-4.1-mini-2025-04-14",
                messages=[
                    {
                        "role": "system", 
                        "content": "あなたは政治的発言を分析し、特定の政策テーマに該当するかを3段階（0/1/2）で判断するアシスタントです。フラグが立つ場合は、根拠となる発言部分を正確に引用し、理由を10-30文字で詳細に説明してください。1つの発言が複数のフラグに該当する場合もあります。必ず完全で有効なJSONで応答してください。"
                    },
                    {"role": "user", "content": prompt}
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "flag_analysis_result",
                        "strict": False,  # enum制限を回避するためstrictモードを無効化
                        "schema": json_schema
                    }
                },
                temperature=self.temp_var.get(),
                max_tokens=8000  # トークン数を倍増
            )
            return response
        except RateLimitError as e:
            # レート制限エラーの詳細解析
            if "tokens" in str(e).lower():
                self.log(f"トークンレート制限エラー: {e}")
                # TPM制限の場合、より長く待機
                if hasattr(e, 'response') and e.response:
                    retry_after = e.response.headers.get('retry-after')
                    if retry_after:
                        wait_time = float(retry_after) + 2  # 追加バッファ
                        self.log(f"推奨待機時間: {wait_time}秒")
                        await asyncio.sleep(wait_time)
            raise
        except APIConnectionError as e:
            self.log(f"API接続エラー: {e}")
            raise
        except APITimeoutError as e:
            self.log(f"APIタイムアウトエラー: {e}")
            raise
        except Exception as e:
            self.log(f"予期しないAPIエラー: {e}")
            self.logger.error(f"Unexpected API error: {e}", exc_info=True)
            raise
    
    async def process_batch(self, client, batch_data, prompt_template, json_schema):
        """バッチを並列処理"""
        tasks = []
        for idx, name, text in batch_data:
            prompt = prompt_template.format(text=text)
            task = self.process_single_item(client, idx, name, text, prompt, json_schema)
            tasks.append(task)
        
        await asyncio.gather(*tasks, return_exceptions=True)
    
    async def process_single_item(self, client, idx, name, text, prompt, json_schema):
        """単一アイテムを処理"""
        try:
            self.log(f"処理中 ({idx+1}): {name}")
            
            # API呼び出し
            response = await self.call_openai_api(client, prompt, json_schema)
            
            # トークン使用量を記録
            if hasattr(response, 'usage') and response.usage:
                input_tokens = getattr(response.usage, 'prompt_tokens', 0)
                output_tokens = getattr(response.usage, 'completion_tokens', 0)
                
                # トークン使用量をデータフレームに追加
                token_row = {
                    "名前": name,
                    "インプットトークン": input_tokens,
                    "アウトプットトークン": output_tokens
                }
                self.token_usage_df = pd.concat([self.token_usage_df, pd.DataFrame([token_row])], ignore_index=True)
                
                self.log(f"  トークン使用量 - 入力: {input_tokens}, 出力: {output_tokens}")
            
            # Structured OutputsのJSONレスポンスを解析
            result_text = response.choices[0].message.content
            
            # JSONバリデーション（強化版）
            try:
                # 不完全なJSONの検出と修復試行
                if result_text and not result_text.strip().endswith('}'):
                    self.log(f"  警告: 不完全なJSON応答を検出 - 修復を試行")
                    # 基本的な修復: 最後に}を追加
                    result_text = result_text.strip()
                    # 開いているブレースをカウントして適切に閉じる
                    open_braces = result_text.count('{') - result_text.count('}')
                    result_text += '}' * max(0, open_braces)
                    
                    # 最後のコンマを削除（JSON仕様に準拠）
                    result_text = result_text.rstrip().rstrip(',')
                    if not result_text.endswith('}'):
                        result_text += '}'
                
                result_dict = json.loads(result_text)
            except json.JSONDecodeError as e:
                # JSONエラーの詳細ログ
                self.log(f"  JSON解析エラー詳細: {e}")
                self.log(f"  問題のあるJSON（最初の500文字）: {result_text[:500]}")
                raise ValueError(f"無効なJSON応答: {e}")
            
            # 結果を行として追加
            result_row = {"名前": name}
            flag_count = 0
            
            for code in self.flags_df["小項目コード"]:
                code_str = str(code)
                if code_str in result_dict:
                    item = result_dict[code_str]
                    
                    # データバリデーション
                    if not isinstance(item, dict):
                        self.log(f"  警告: コード{code_str}の結果が辞書型ではありません: {type(item)}")
                        # デフォルト値を設定
                        result_row[f"{code_str}_flag"] = 0
                        result_row[f"{code_str}_quote"] = ""
                        result_row[f"{code_str}_reason"] = ""
                        continue
                    
                    flag = item.get("flag", 0)
                    quote = item.get("quote", "")
                    reason = item.get("reason", "")
                    
                    # 文字列の清浄化（引用符などの問題文字を除去）
                    if isinstance(quote, str):
                        quote = quote.replace('"', "'").replace('\n', ' ').strip()
                    if isinstance(reason, str):
                        reason = reason.replace('"', "'").replace('\n', ' ').strip()
                    
                    # フラグ値の検証
                    if flag not in [0, 1, 2]:
                        self.log(f"  警告: 不正なフラグ値 {flag} を 0 に修正")
                        flag = 0
                    
                    # フラグに基づいて引用と理由を設定
                    if flag in [1, 2]:
                        result_row[f"{code_str}_flag"] = flag
                        result_row[f"{code_str}_quote"] = quote
                        result_row[f"{code_str}_reason"] = reason
                        flag_count += 1
                    else:
                        # フラグが0の場合でも明示的に設定
                        result_row[f"{code_str}_flag"] = 0
                        result_row[f"{code_str}_quote"] = ""
                        result_row[f"{code_str}_reason"] = ""
                else:
                    # デフォルト値
                    result_row[f"{code_str}_flag"] = 0
                    result_row[f"{code_str}_quote"] = ""
                    result_row[f"{code_str}_reason"] = ""
            
            # スレッドセーフな方法で結果を追加
            await self.add_result_safely(result_row)
            self.processed_indices.add(idx)
            
            # 失敗リストから削除（再処理成功時）
            if idx in self.failed_indices:
                self.failed_indices.remove(idx)
                self.log(f"  再処理成功: {name}")
            
            self.log(f"  完了: {flag_count}個のフラグが立てられました")
            
        except tenacity.RetryError as e:
            error_msg = f"最大リトライ回数に到達 - {name}: {str(e.last_attempt.exception())}"
            self.log(f"  {error_msg}")
            self.failed_indices.add(idx)
            self.error_log.append(f"[{datetime.now()}] {error_msg}")
            self.logger.error(error_msg, exc_info=True)
            
        except (APIError, ValueError, json.JSONDecodeError) as e:
            error_msg = f"処理エラー - {name}: {str(e)}"
            self.log(f"  {error_msg}")
            self.failed_indices.add(idx)
            self.error_log.append(f"[{datetime.now()}] {error_msg}")
            self.logger.error(error_msg, exc_info=True)
            
        except Exception as e:
            error_msg = f"予期しないエラー - {name}: {str(e)}"
            self.log(f"  {error_msg}")
            self.failed_indices.add(idx)
            self.error_log.append(f"[{datetime.now()}] {error_msg}")
            self.logger.error(error_msg, exc_info=True)
    
    async def add_result_safely(self, result_row):
        """スレッドセーフに結果を追加"""
        self.results_df = pd.concat([self.results_df, pd.DataFrame([result_row])], ignore_index=True)
    
    def display_results(self):
        """結果を表示"""
        if self.results_df is None or self.results_df.empty:
            return
        
        # 既存の項目をクリア
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        
        # カラムを設定
        display_columns = ["名前"]
        for code in self.flags_df["小項目コード"]:
            code_str = str(code)
            flag_info = self.flags_df[self.flags_df["小項目コード"] == int(code_str)]
            if not flag_info.empty:
                item_name = flag_info.iloc[0]['小項目名']
                display_columns.append(f"{code_str}: {item_name}")
        
        self.result_tree["columns"] = display_columns
        
        # ヘッダーを設定
        self.result_tree.heading("#0", text="")
        self.result_tree.column("#0", width=0, stretch=tk.NO)
        
        for col in display_columns:
            self.result_tree.heading(col, text=col)
            if col == "名前":
                self.result_tree.column(col, width=100, stretch=tk.YES)
            else:
                self.result_tree.column(col, width=80, stretch=tk.YES)
        
        # データ行を追加
        for i, row in self.results_df.iterrows():
            values = [row["名前"]]
            for code in self.flags_df["小項目コード"]:
                code_str = str(code)
                flag = row.get(f"{code_str}_flag", 0)
                values.append(str(flag))
            
            self.result_tree.insert("", tk.END, text="", values=values)
    
    def save_results(self):
        """結果を保存"""
        if self.results_df is None or self.results_df.empty:
            messagebox.showwarning("警告", "保存する結果がありません")
            return
        
        output_format = self.output_format_var.get()
        
        if output_format == "excel":
            self.save_excel_results()
        else:
            self.save_csv_results()
    
    def save_excel_results(self):
        """Excel形式で結果を保存（4つのシート：フラグ・引用・理由・トークン）"""
        file_path = filedialog.asksaveasfilename(
            title="結果を保存",
            defaultextension=".xlsx",
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            # データフレームを準備
            flag_df = pd.DataFrame()
            quote_df = pd.DataFrame()
            reason_df = pd.DataFrame()
            
            # ヘッダー情報を準備
            header_row1 = ["名前"]
            header_row2 = [""]
            column_names = ["名前"]
            
            for code in self.flags_df["小項目コード"]:
                code_str = str(code)
                flag_info = self.flags_df[self.flags_df["小項目コード"] == int(code_str)]
                
                if not flag_info.empty:
                    category = str(flag_info.iloc[0]['大項目'])
                    item_name = str(flag_info.iloc[0]['小項目名'])
                    description = str(flag_info.iloc[0]['説明'])
                    
                    col_name = f"{category}-{item_name}"
                    column_names.append(col_name)
                    header_row1.append(col_name)
                    header_row2.append(description)
                else:
                    column_names.append(code_str)
                    header_row1.append(code_str)
                    header_row2.append("")
            
            # データ行を準備
            flag_data = []
            quote_data = []
            reason_data = []
            
            for _, row in self.results_df.iterrows():
                name = row["名前"]
                flag_row = [name]
                quote_row = [name]
                reason_row = [name]
                
                for code in self.flags_df["小項目コード"]:
                    code_str = str(code)
                    flag = row.get(f"{code_str}_flag", 0)
                    quote = row.get(f"{code_str}_quote", "")
                    reason = row.get(f"{code_str}_reason", "")
                    
                    # 数値型として明示的に保存
                    flag_row.append(int(flag) if flag != "" else 0)
                    quote_row.append(str(quote))
                    reason_row.append(str(reason))
                
                flag_data.append(flag_row)
                quote_data.append(quote_row)
                reason_data.append(reason_row)
            
            # DataFrameを作成
            flag_df = pd.DataFrame(flag_data, columns=column_names)
            quote_df = pd.DataFrame(quote_data, columns=column_names)
            reason_df = pd.DataFrame(reason_data, columns=column_names)
            
            # トークンシート用データを準備
            token_df = self.token_usage_df.copy()
            
            # 料金計算を追加（安全な処理）
            if not token_df.empty and 'インプットトークン' in token_df.columns and 'アウトプットトークン' in token_df.columns:
                # 安全な数値変換とエラーハンドリング
                try:
                    input_tokens = pd.to_numeric(token_df['インプットトークン'], errors='coerce').fillna(0)
                    output_tokens = pd.to_numeric(token_df['アウトプットトークン'], errors='coerce').fillna(0)
                    
                    # 料金計算（入力$0.40/1M、出力$1.60/1M）
                    input_costs = input_tokens * 0.40 / 1000000
                    output_costs = output_tokens * 1.60 / 1000000
                    total_costs = input_costs + output_costs
                    
                    # DataFrameに追加（既にfloat型なのでそのまま）
                    token_df['入力料金($)'] = input_costs.round(4)
                    token_df['出力料金($)'] = output_costs.round(4)
                    token_df['合計料金($)'] = total_costs.round(4)
                    
                except Exception as e:
                    self.log(f"料金計算エラー: {e}")
                    # エラーが発生した場合はデフォルト値を設定
                    token_df['入力料金($)'] = 0.0
                    token_df['出力料金($)'] = 0.0
                    token_df['合計料金($)'] = 0.0
            
            # Excelファイルに書き込む
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # フラグシート
                flag_df.to_excel(writer, sheet_name='フラグ', index=False)
                # 引用シート
                quote_df.to_excel(writer, sheet_name='引用', index=False)
                # 理由シート
                reason_df.to_excel(writer, sheet_name='理由', index=False)
                # トークンシート
                token_df.to_excel(writer, sheet_name='トークン', index=False)
                
                # フラグ・引用・理由シートに説明行を追加
                workbook = writer.book
                for sheet_name in ['フラグ', '引用', '理由']:
                    worksheet = workbook[sheet_name]
                    worksheet.insert_rows(2)
                    for i, desc in enumerate(header_row2):
                        worksheet.cell(row=2, column=i+1, value=desc)
                
                # トークンシートの合計行を追加
                if not token_df.empty:
                    token_sheet = workbook['トークン']
                    # 最終行を取得
                    last_row = token_sheet.max_row + 1
                    
                    # 合計計算（安全な数値変換）
                    input_sum = pd.to_numeric(token_df['インプットトークン'], errors='coerce').sum()
                    output_sum = pd.to_numeric(token_df['アウトプットトークン'], errors='coerce').sum()
                    
                    # 合計行を追加
                    token_sheet.cell(row=last_row, column=1, value="合計")
                    token_sheet.cell(row=last_row, column=2, value=int(input_sum))
                    token_sheet.cell(row=last_row, column=3, value=int(output_sum))
                    
                    # 料金列が存在する場合のみ合計を計算
                    if len(token_df.columns) > 3:
                        # 料金の合計を直接計算（roundを使わない）
                        input_cost_sum = float(input_sum * 0.40 / 1000000)
                        output_cost_sum = float(output_sum * 1.60 / 1000000)
                        total_cost_sum = input_cost_sum + output_cost_sum
                        
                        # セルに値を設定（事前にroundした値を使用）
                        token_sheet.cell(row=last_row, column=4, value=round(input_cost_sum, 4))
                        token_sheet.cell(row=last_row, column=5, value=round(output_cost_sum, 4))
                        token_sheet.cell(row=last_row, column=6, value=round(total_cost_sum, 4))
            
            self.log(f"結果をExcelに保存しました: {file_path}")
            
            # 最終サマリーをログに出力（安全な処理）
            if not token_df.empty:
                try:
                    total_input = pd.to_numeric(token_df['インプットトークン'], errors='coerce').sum()
                    total_output = pd.to_numeric(token_df['アウトプットトークン'], errors='coerce').sum()
                    total_cost = (total_input * 0.40 + total_output * 1.60) / 1000000
                    
                    # 数値フォーマット
                    total_input_str = f"{int(total_input):,}"
                    total_output_str = f"{int(total_output):,}"
                    total_cost_str = f"{total_cost:.4f}"
                    
                    self.log(f"トークン合計 - 入力: {total_input_str}, 出力: {total_output_str}, 料金: ${total_cost_str}")
                except Exception as e:
                    self.log(f"トークン合計計算エラー: {e}")
                    self.log("トークン合計 - 計算できませんでした")
            
            messagebox.showinfo("保存完了", f"結果をExcelに保存しました: {file_path}")
            
        except Exception as e:
            self.log(f"Excel保存エラー: {e}")
            self.log(f"詳細: {traceback.format_exc()}")
            messagebox.showerror("保存エラー", f"結果の保存中にエラーが発生しました: {e}")
    
    def save_csv_results(self):
        """CSV形式で結果を保存"""
        file_path = filedialog.asksaveasfilename(
            title="結果を保存",
            defaultextension=".csv",
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            # 結果をそのまま保存
            self.results_df.to_csv(file_path, index=False, encoding='utf-8-sig')
            self.log(f"結果をCSVに保存しました: {file_path}")
            messagebox.showinfo("保存完了", f"結果をCSVに保存しました: {file_path}")
            
        except Exception as e:
            self.log(f"CSV保存エラー: {e}")
            self.log(f"詳細: {traceback.format_exc()}")
            messagebox.showerror("保存エラー", f"結果の保存中にエラーが発生しました: {e}")

# アプリケーション起動
if __name__ == "__main__":
    root = tk.Tk()
    app = FlagAnalysisApp(root)
    root.mainloop()
